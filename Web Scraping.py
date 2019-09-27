
# coding: utf-8

# # Importing the required Libraries

# In[1]:


from bs4 import BeautifulSoup as soup 
from urllib.request import urlopen as uReq


# In[2]:


url = 'https://www.olx.in/items/q-maruti-Suzuki-cars'


# # Accesing Data from OLX

# In[3]:


Client = uReq(url)
page_html = Client.read()
page_soup = soup(page_html,'html.parser')
names = page_soup.findAll('span',{'class':'_2tW1I'})
years = page_soup.findAll('span',{'class':"_2TVI3"})
price = page_soup.findAll('span',{'class':'_89yzn'})
len(price)


# # Storing the Data into Lists and Cleaning It

# In[4]:


name_list = []
s = []
p=[]
price_list=[]
kms_list = []
year_list = [] 
for i in range(len(names)):
    name_list.append(names[i].text)
    s.append(years[i].text)
    p.append(price[i].text)


# In[5]:


for char in s:
    for i in range(len(char)):
        if  char[i]== '-':
            kms_list.append(int(char[i+1:(len(char)-4)]))
            year_list.append(int(char[0:i-1]))


# In[6]:


name_list


# In[7]:


kms_list


# In[8]:


year_list


# In[9]:


p   # this list has comma and rupee symbol - so can't be converted to 


# In[10]:


for char in p:
    for i in range(len(char)):
        if char[i] == '₹':
            price_list.append((char[i+2:len(char)]))


# In[11]:


price = []
for num in price_list:
    n=''
    for i in num:
        if i!=',':
            n+=i
    price.append(int(n))
price_list = price


# In[12]:


# Converting the year to number of years old the car is 
# Used numpy because it is simple for arithmetic operations

import numpy as np
years = np.array(year_list)
print(years)
years = 2019 - years
year_list = list(years)


# # Storing all the Collected Data into an Excel File

# In[13]:


import openpyxl


# In[14]:


f = r"E:\Projects\Python Project\Car Resale Price Predictor\Car Resale.xlsx"


# In[15]:


wkbook = openpyxl.load_workbook(f)


# In[16]:


sheet = wkbook['Sheet1']


# In[17]:


sheet.cell(row = 1,column=1).value = 'Car Name'
sheet.cell(row = 1,column=2).value = 'Years'
sheet.cell(row = 1,column=3).value = 'Running'
sheet.cell(row = 1,column=4).value = 'Resale Price'


# In[18]:


for i in range(20):
    sheet.cell(row= i+2 , column =1).value = name_list[i]
    sheet.cell(row=i+2  , column =2).value = year_list[i]
    sheet.cell(row=i+2  , column =3).value = kms_list[i]
    sheet.cell(row= i+2 , column =4).value = price_list[i]
    wkbook.save(r'E:\Projects\Python Project\Car Resale Price Predictor\Car Resale Data.csv') 

